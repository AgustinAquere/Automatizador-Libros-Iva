from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import os
import io
import pickle
from typing import List, Dict, Optional


class DriveHandler:
    """Maneja todas las operaciones con Google Drive"""
    
    SCOPES = ['https://www.googleapis.com/auth/drive']
    ROOT_FOLDER_NAME = "Clientes Libros Iva"  # Carpeta de prueba
    
    def __init__(self, credentials_path: str = 'credentials.json', token_path: str = 'token.pickle'):
        self.credentials_path = credentials_path
        self.token_path = token_path
        self.service = None
        self.root_folder_id = None
        
    def authenticate(self):
        """Autentica con Google Drive usando OAuth2"""
        creds = None
        
        # El archivo token.pickle almacena los tokens de acceso del usuario
        if os.path.exists(self.token_path):
            try:
                with open(self.token_path, 'rb') as token:
                    creds = pickle.load(token)
            except Exception as e:
                print(f"⚠️  Error al cargar token: {str(e)}")
                print(f"   Eliminando token corrupto...")
                os.unlink(self.token_path)
                creds = None
        
        # Si no hay credenciales válidas, hacer login
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    print("🔄 Refrescando token expirado...")
                    creds.refresh(Request())
                    print("✓ Token refrescado exitosamente")
                except Exception as e:
                    print(f"⚠️  Token expirado o revocado: {str(e)}")
                    print(f"   Eliminando token y solicitando nueva autenticación...")
                    if os.path.exists(self.token_path):
                        os.unlink(self.token_path)
                    creds = None
            
            if not creds:
                print("🔐 Iniciando proceso de autenticación...")
                print("   Se abrirá tu navegador para autorizar el acceso...")
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_path, self.SCOPES)
                creds = flow.run_local_server(port=0)
                print("✓ Autenticación exitosa")
            
            # Guardar las credenciales para la próxima ejecución
            try:
                with open(self.token_path, 'wb') as token:
                    pickle.dump(creds, token)
                print(f"✓ Token guardado en {self.token_path}")
            except Exception as e:
                print(f"⚠️  No se pudo guardar el token: {str(e)}")
        
        self.service = build('drive', 'v3', credentials=creds)
        return self.service
    
    def find_folder(self, folder_name: str, parent_id: Optional[str] = None) -> Optional[str]:
        """
        Busca una carpeta por nombre
        Returns: ID de la carpeta o None si no existe
        """
        if not self.service:
            self.authenticate()
        
        query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        if parent_id:
            query += f" and '{parent_id}' in parents"
        
        results = self.service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name)'
        ).execute()
        
        items = results.get('files', [])
        return items[0]['id'] if items else None
    
    def get_root_folder_id(self) -> str:
        """Obtiene el ID de la carpeta raíz 'Clientes libros iva'"""
        if self.root_folder_id:
            return self.root_folder_id
        
        self.root_folder_id = self.find_folder(self.ROOT_FOLDER_NAME)
        if not self.root_folder_id:
            raise Exception(f"No se encontró la carpeta '{self.ROOT_FOLDER_NAME}' en Drive")
        
        return self.root_folder_id
    
    def list_clients(self) -> List[Dict]:
        """
        Lista todos los clientes (carpetas dentro de 'Clientes libros iva')
        Returns: Lista de dicts con {name, id, enabled}
        """
        if not self.service:
            self.authenticate()
        
        root_id = self.get_root_folder_id()
        
        query = f"'{root_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
        
        results = self.service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name)',
            orderBy='name'
        ).execute()
        
        items = results.get('files', [])
        
        # Filtrar el archivo "cuits" y agregar estado enabled
        clients = []
        for item in items:
            if item['name'].lower() == 'cuits':
                continue
            
            clients.append({
                'name': item['name'],
                'id': item['id'],
                'enabled': item['name'].lower() not in self.DISABLED_CLIENTS
            })
        
        return clients
    
    def get_client_structure(self, client_name: str) -> Dict:
        """
        Obtiene la estructura de carpetas de un cliente
        Returns: {ventas_id, compras_id}
        """
        if not self.service:
            self.authenticate()
        
        root_id = self.get_root_folder_id()
        client_id = self.find_folder(client_name, root_id)
        
        if not client_id:
            raise Exception(f"No se encontró el cliente '{client_name}'")
        
        ventas_id = self.find_folder("Ventas", client_id)
        compras_id = self.find_folder("Compras", client_id)
        
        return {
            'client_id': client_id,
            'ventas_id': ventas_id,
            'compras_id': compras_id
        }

    def create_year_file(self, client_name: str, tipo: str, year: int) -> str:
        """
        Crea un nuevo archivo Excel vacío para el año
        Returns: file_id del archivo creado
        """
        import openpyxl
        import tempfile

        print(f"   📝 Obteniendo estructura del cliente '{client_name}'...")
        structure = self.get_client_structure(client_name)
        folder_id = structure['ventas_id'] if tipo.lower() == 'ventas' else structure['compras_id']

        if not folder_id:
            raise Exception(f"No se encontró la carpeta '{tipo}' para el cliente '{client_name}'")

        print(f"   📝 Carpeta {tipo}: {folder_id}")

        # Formato: "Libro IVA Ventas 2025 Cliente"
        filename = f"Libro Iva {tipo.capitalize()} {year} {client_name}.xlsx"

        # Crear un Excel vacío temporal
        print(f"   📝 Creando archivo Excel vacío...")
        wb = openpyxl.Workbook()

        # Eliminar TODAS las hojas por defecto
        for sheet in wb.worksheets:
            wb.remove(sheet)

        # Crear una hoja temporal (se eliminará cuando se agregue la primera pestaña real)
        ws = wb.create_sheet(title="_temp")
        ws['A1'] = "Archivo temporal - Esta pestaña se eliminará al agregar datos"

        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp_path = tmp.name

        print(f"   📝 Guardando en temporal: {tmp_path}")
        wb.save(tmp_path)

        # Subir a Drive
        print(f"   📝 Subiendo '{filename}' a Drive...")
        file_id = self.upload_file(tmp_path, folder_id, filename)

        # Limpiar archivo temporal
        import os
        os.unlink(tmp_path)

        print(f"   ✅ Archivo creado exitosamente con ID: {file_id}")

        return file_id

    def check_year_file_exists(self, client_name: str, tipo: str, year: int) -> Dict:
        """
        Verifica si existe el archivo del año
        Returns: {exists: bool, file_id: str or None}
        """
        structure = self.get_client_structure(client_name)
        folder_id = structure['ventas_id'] if tipo.lower() == 'ventas' else structure['compras_id']
        
        if not folder_id:
            raise Exception(f"No se encontró la carpeta '{tipo}' para el cliente '{client_name}'")
        
        # Formato: "Libro IVA Ventas 2025 Cliente"
        filename = f"Libro Iva {tipo.capitalize()} {year} {client_name}.xlsx"
        
        print(f"   🔎 Buscando archivo: '{filename}'")
        print(f"   🔎 En carpeta: {folder_id}")
        
        # Escapar comillas simples en el nombre del archivo para la query
        filename_escaped = filename.replace("'", "\\'")
        query = f"name='{filename_escaped}' and '{folder_id}' in parents and trashed=false"
        
        print(f"   🔎 Query: {query}")
        
        results = self.service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name)'
        ).execute()
        
        items = results.get('files', [])
        
        print(f"   🔎 Archivos encontrados: {len(items)}")
        if items:
            for item in items:
                print(f"      - {item['name']} (ID: {item['id']})")
        
        return {
            'exists': len(items) > 0,
            'file_id': items[0]['id'] if items else None,
            'folder_id': folder_id,
            'filename': filename
        }
    
    def download_file(self, file_id: str, output_path: str):
        """Descarga un archivo de Drive"""
        if not self.service:
            self.authenticate()
        
        # Primero verificar el tipo de archivo
        file_metadata = self.service.files().get(fileId=file_id, fields='mimeType, name').execute()
        mime_type = file_metadata.get('mimeType')
        
        print(f"      🔹 Tipo de archivo: {mime_type}")
        
        # Si es Google Sheets, exportar como Excel
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            print(f"      🔹 Exportando Google Sheets como Excel...")
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Es un archivo Excel normal
            print(f"      🔹 Descargando archivo Excel...")
            request = self.service.files().get_media(fileId=file_id)
        
        with io.FileIO(output_path, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                if status:
                    print(f"      🔹 Progreso: {int(status.progress() * 100)}%")
    
    def upload_file(self, file_path: str, folder_id: str, file_name: str) -> str:
        """
        Sube un archivo a Drive como Excel (NO como Google Sheets)
        Returns: ID del archivo subido
        """
        if not self.service:
            self.authenticate()
        
        print(f"      🔹 Preparando subida de '{file_name}' a carpeta {folder_id}")
        
        file_metadata = {
            'name': file_name,
            'parents': [folder_id],
            # NO incluir mimeType en metadata para que mantenga el formato Excel
        }
        
        media = MediaFileUpload(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        print(f"      🔹 Subiendo archivo como Excel binario (NO Google Sheets)...")
        file = self.service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, name, mimeType, parents',
            supportsAllDrives=True
        ).execute()
        
        file_id = file.get('id')
        mime_type = file.get('mimeType')
        print(f"      ✅ Archivo subido: {file.get('name')}")
        print(f"      ✅ Tipo MIME: {mime_type}")
        print(f"      ✅ ID: {file_id}")
        
        return file_id

    def update_file(self, file_id: str, file_path: str):
        """Actualiza un archivo existente en Drive"""
        if not self.service:
            self.authenticate()

        # Verificar tipo de archivo actual
        file_metadata = self.service.files().get(fileId=file_id, fields='mimeType, name').execute()
        current_mime = file_metadata.get('mimeType')

        print(f"      🔹 Actualizando archivo {file_id}...")
        print(f"      🔹 Tipo actual: {current_mime}")

        # Si el archivo actual es Google Sheets, necesitamos eliminarlo y crear uno nuevo
        if current_mime == 'application/vnd.google-apps.spreadsheet':
            print(f"      ⚠️  Archivo actual es Google Sheets, será reemplazado por Excel...")
            # Obtener info del archivo
            file_info = self.service.files().get(fileId=file_id, fields='name, parents').execute()
            file_name = file_info.get('name')
            parents = file_info.get('parents', [])

            # Eliminar el Google Sheets
            self.service.files().delete(fileId=file_id).execute()
            print(f"      🔹 Google Sheets eliminado")

            # Subir el nuevo Excel
            if parents:
                new_file_id = self.upload_file(file_path, parents[0], file_name)
                print(f"      ✅ Nuevo archivo Excel creado con ID: {new_file_id}")
                return new_file_id
        else:
            # Es Excel, actualizar normalmente
            media = MediaFileUpload(
                file_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                resumable=True
            )

            updated_file = self.service.files().update(
                fileId=file_id,
                media_body=media,
                fields='id, name, mimeType, modifiedTime'
            ).execute()

            print(f"      ✅ Archivo actualizado: {updated_file.get('name')}")
            print(f"      ✅ Tipo: {updated_file.get('mimeType')}")
            print(f"      ✅ Modificado: {updated_file.get('modifiedTime')}")
            return file_id
    
    def create_client(self, client_name: str, cuit: str) -> Dict:
        """
        Crea un nuevo cliente en Drive con su estructura de carpetas
        Returns: {success: bool, client_id: str, message: str}
        """
        if not self.service:
            self.authenticate()
        
        root_id = self.get_root_folder_id()
        
        print(f"   📝 Creando cliente '{client_name}' (CUIT: {cuit})...")
        
        # Verificar si ya existe
        existing = self.find_folder(client_name, root_id)
        if existing:
            return {
                'success': False,
                'message': f"El cliente '{client_name}' ya existe"
            }
        
        # Crear carpeta del cliente
        client_metadata = {
            'name': client_name,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [root_id]
        }
        
        client_folder = self.service.files().create(
            body=client_metadata,
            fields='id, name'
        ).execute()
        
        client_id = client_folder.get('id')
        print(f"   ✓ Carpeta cliente creada: {client_id}")
        
        # Crear carpeta Ventas
        ventas_metadata = {
            'name': 'Ventas',
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [client_id]
        }
        
        ventas_folder = self.service.files().create(
            body=ventas_metadata,
            fields='id, name'
        ).execute()
        
        print(f"   ✓ Carpeta Ventas creada: {ventas_folder.get('id')}")
        
        # Crear carpeta Compras
        compras_metadata = {
            'name': 'Compras',
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [client_id]
        }
        
        compras_folder = self.service.files().create(
            body=compras_metadata,
            fields='id, name'
        ).execute()
        
        print(f"   ✓ Carpeta Compras creada: {compras_folder.get('id')}")
        
        return {
            'success': True,
            'client_id': client_id,
            'ventas_id': ventas_folder.get('id'),
            'compras_id': compras_folder.get('id'),
            'message': f"Cliente '{client_name}' creado exitosamente"
        }
    
    def find_client_by_cuit(self, cuit: str) -> Dict:
        """
        Busca un cliente por su CUIT
        Returns: {found: bool, client_name: str or None}
        """
        # Por ahora buscaremos en el nombre de las carpetas
        # Podrías mejorar esto guardando los CUIT en algún archivo de metadata
        clients = self.list_clients()
        
        # Buscar si algún cliente tiene el CUIT en su nombre o descripción
        # Por simplicidad, retornamos None por ahora
        # La idea es que guardes un mapeo CUIT -> Nombre en algún lugar
        
        return {
            'found': False,
            'client_name': None,
            'cuit': cuit
        }
        """
        Crea un nuevo archivo Excel vacío para el año
        Returns: file_id del archivo creado
        """
        import openpyxl
        import tempfile
        
        print(f"   📝 Obteniendo estructura del cliente '{client_name}'...")
        structure = self.get_client_structure(client_name)
        folder_id = structure['ventas_id'] if tipo.lower() == 'ventas' else structure['compras_id']
        
        if not folder_id:
            raise Exception(f"No se encontró la carpeta '{tipo}' para el cliente '{client_name}'")
        
        print(f"   📝 Carpeta {tipo}: {folder_id}")
        
        # Formato: "Libro IVA Ventas 2025 Cliente"
        filename = f"Libro Iva {tipo.capitalize()} {year} {client_name}.xlsx"
        
        # Crear un Excel vacío temporal
        print(f"   📝 Creando archivo Excel vacío...")
        wb = openpyxl.Workbook()
        
        # Eliminar TODAS las hojas por defecto
        for sheet in wb.worksheets:
            wb.remove(sheet)
        
        # Crear una hoja temporal (se eliminará cuando se agregue la primera pestaña real)
        ws = wb.create_sheet(title="_temp")
        ws['A1'] = "Archivo temporal - Esta pestaña se eliminará al agregar datos"
        
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp_path = tmp.name
        
        print(f"   📝 Guardando en temporal: {tmp_path}")
        wb.save(tmp_path)
        
        # Subir a Drive
        print(f"   📝 Subiendo '{filename}' a Drive...")
        file_id = self.upload_file(tmp_path, folder_id, filename)
        
        # Limpiar archivo temporal
        import os
        os.unlink(tmp_path)
        
        print(f"   ✅ Archivo creado exitosamente con ID: {file_id}")
        
        return file_id


def test_connection():
    """Función de prueba para verificar la conexión"""
    handler = DriveHandler()
    handler.authenticate()
    
    try:
        clients = handler.list_clients()
        print(f"✓ Conexión exitosa. Se encontraron {len(clients)} clientes:")
        for client in clients:
            status = "✓" if client['enabled'] else "✗ (deshabilitado)"
            print(f"  {status} {client['name']}")
        return True
    except Exception as e:
        print(f"✗ Error: {str(e)}")
        return False
