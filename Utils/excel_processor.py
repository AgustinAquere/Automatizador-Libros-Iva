import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment


class ExcelProcessor:
    """Procesa y limpia libros de IVA"""
    
    # Columnas que siempre se mantienen antes de "Moneda"
    BASE_COLUMNS = [
        "Fecha", "Tipo", "Punto de Venta", "Número Desde",
        "Nro. Doc. Receptor", "Denominación Receptor", 
        "Tipo Cambio", "Moneda"
    ]
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.df = None
        self.month_detected = None
        
    def read_excel(self) -> pd.DataFrame:
        """Lee el Excel y retorna el DataFrame"""
        # SIEMPRE saltar la primera fila (contiene el CUIT/título)
        # Las columnas están en la segunda fila
        self.df = pd.read_excel(self.file_path, skiprows=1)
        
        return self.df
    
    def detect_info_from_header(self) -> Dict:
        """
        Detecta información del encabezado (primera fila antes de skiprows)
        Returns: {cuit: str, tipo: str (ventas/compras)}
        """
        import re
        
        # Leer SOLO la primera fila, sin procesar
        df_header = pd.read_excel(self.file_path, nrows=1, header=None)
        
        # La primera celda (columna 0, fila 0) contiene algo como:
        # "Mis Comprobantes Emitidos - CUIT 30716820080"
        # "Mis Comprobantes Recibidos - CUIT 30716820080"
        
        header_text = str(df_header.iloc[0, 0])
        
        print(f"   🔍 Analizando encabezado: {header_text}")
        
        # Si no contiene la info esperada, intentar con otras celdas de la primera fila
        if 'comprobantes' not in header_text.lower() and 'cuit' not in header_text.lower():
            # Buscar en otras columnas de la primera fila
            for col in range(min(5, len(df_header.columns))):
                cell_value = str(df_header.iloc[0, col])
                if 'comprobantes' in cell_value.lower() or 'cuit' in cell_value.lower():
                    header_text = cell_value
                    print(f"   🔍 Encontrado en columna {col}: {header_text}")
                    break
        
        # Detectar tipo (Emitidos = Ventas, Recibidos = Compras)
        tipo = None
        if 'emitidos' in header_text.lower():
            tipo = 'ventas'
        elif 'recibidos' in header_text.lower():
            tipo = 'compras'
        
        # Extraer CUIT usando regex (busca 11 dígitos consecutivos después de "CUIT")
        cuit_match = re.search(r'CUIT\s*[:\-]?\s*(\d{11})', header_text, re.IGNORECASE)
        cuit = cuit_match.group(1) if cuit_match else None
        
        # Si no encontró con "CUIT", buscar cualquier secuencia de 11 dígitos
        if not cuit:
            cuit_match = re.search(r'\b(\d{11})\b', header_text)
            cuit = cuit_match.group(1) if cuit_match else None
        
        print(f"   ✓ CUIT detectado: {cuit}")
        print(f"   ✓ Tipo detectado: {tipo}")
        
        if not cuit or not tipo:
            raise ValueError(
                f"No se pudo detectar CUIT o tipo en el encabezado.\n"
                f"Encabezado encontrado: '{header_text}'\n"
                f"Asegúrese de que la primera fila contenga algo como:\n"
                f"'Mis Comprobantes Emitidos - CUIT 30716820080'"
            )
        
        return {
            'cuit': cuit,
            'tipo': tipo
        }
        """
        Detecta el mes y año predominante en el Excel
        Returns: (mes, año)
        """
        if self.df is None:
            self.read_excel()
        
        # Hacer una copia para no modificar el original
        df_temp = self.df.copy()
        
        # Convertir la columna Fecha a datetime si no lo es ya
        if df_temp['Fecha'].dtype == 'object':
            # Intentar varios formatos comunes
            df_temp['Fecha'] = pd.to_datetime(df_temp['Fecha'], format='%d/%m/%Y', errors='coerce')
        
        # Eliminar filas con fechas inválidas
        df_temp = df_temp.dropna(subset=['Fecha'])
        
        if len(df_temp) == 0:
            raise ValueError("No se encontraron fechas válidas en el archivo")
        
        # Encontrar el mes más frecuente
        month_counts = df_temp['Fecha'].dt.month.value_counts()
        most_common_month = month_counts.idxmax()
        
        # Obtener el año correspondiente
        year_series = df_temp[df_temp['Fecha'].dt.month == most_common_month]['Fecha'].dt.year
        year = int(year_series.mode()[0]) if len(year_series) > 0 else int(year_series.iloc[0])
        
        self.month_detected = (int(most_common_month), year)
        return int(most_common_month), year

    def detect_month(self) -> Tuple[int, int]:
        """
        Detecta el mes y año predominante en el Excel
        Returns: (mes, año)
        """
        if self.df is None:
            self.read_excel()

        # Hacer una copia para no modificar el original
        df_temp = self.df.copy()

        # Convertir la columna Fecha a datetime si no lo es ya
        if df_temp['Fecha'].dtype == 'object':
            # Intentar varios formatos comunes
            df_temp['Fecha'] = pd.to_datetime(df_temp['Fecha'], format='%d/%m/%Y', errors='coerce')

        # Eliminar filas con fechas inválidas
        df_temp = df_temp.dropna(subset=['Fecha'])

        if len(df_temp) == 0:
            raise ValueError("No se encontraron fechas válidas en el archivo")

        # Encontrar el mes más frecuente
        # Le indicamos que el DÍA va primero (formato DD/MM/AAAA de AFIP)
        df_temp['Fecha'] = pd.to_datetime(df_temp['Fecha'], dayfirst=True, errors='coerce')

        # Ahora el conteo de meses va a ser el correcto
        month_counts = df_temp['Fecha'].dt.month.value_counts()
        most_common_month = month_counts.idxmax()

        # Obtener el año correspondiente
        year_series = df_temp[df_temp['Fecha'].dt.month == most_common_month]['Fecha'].dt.year
        year = int(year_series.mode()[0]) if len(year_series) > 0 else int(year_series.iloc[0])

        self.month_detected = (int(most_common_month), year)
        return int(most_common_month), year
    
    def get_month_name(self, month: int) -> str:
        """Convierte número de mes a nombre en español"""
        months = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        return months.get(month, str(month))
    
    def clean_data(self) -> pd.DataFrame:
        """
        Limpia el DataFrame:
        1. Elimina columnas vacías después de 'Moneda'
        2. Multiplica montos por tipo de cambio
        3. Convierte Notas de Crédito a negativo
        4. Agrega fila de totales
        """
        if self.df is None:
            self.read_excel()
        
        # Hacer una copia para trabajar
        df_clean = self.df.copy()
        
        # Encontrar el índice de la columna "Moneda"
        try:
            moneda_idx = df_clean.columns.get_loc("Moneda")
        except KeyError:
            raise ValueError("No se encontró la columna 'Moneda' en el Excel")
        
        # Obtener todas las columnas después de "Moneda"
        columns_after_moneda = df_clean.columns[moneda_idx + 1:]
        
        # Identificar columnas a eliminar (las que están completamente vacías)
        columns_to_drop = []
        for col in columns_after_moneda:
            if df_clean[col].isna().all() or (df_clean[col] == 0).all():
                columns_to_drop.append(col)
        
        # Eliminar columnas vacías
        df_clean = df_clean.drop(columns=columns_to_drop)
        
        # Identificar columnas numéricas (montos) después de "Moneda"
        remaining_cols = [col for col in df_clean.columns[moneda_idx + 1:]]
        numeric_cols = df_clean[remaining_cols].select_dtypes(include=['float64', 'int64']).columns.tolist()
        
        # Identificar notas de crédito
        # Busca en la columna "Tipo" patrones como:
        # - "3 - Nota de Crédito A"
        # - "Nota de Credito"
        # - "NC"
        # etc. (case insensitive)
        is_nota_credito = df_clean['Tipo'].astype(str).str.contains('nota de cr[eé]dito|nc', case=False, na=False)
        
        # Multiplicar cada columna numérica por el tipo de cambio
        # Y convertir a negativo si es nota de crédito
        for col in numeric_cols:
            # Primero multiplicar por tipo de cambio
            df_clean[col] = df_clean[col] * df_clean['Tipo Cambio']
            
            # Luego convertir a negativo si es nota de crédito y el valor es positivo
            # IMPORTANTE: Si el valor ya viene negativo del sistema, no lo tocamos
            # Solo convertimos los positivos a negativos
            df_clean.loc[is_nota_credito & (df_clean[col] > 0), col] = -df_clean.loc[is_nota_credito & (df_clean[col] > 0), col]
        
        # Formatear la columna Fecha si es necesario
        # Verificar si ya es string o si es datetime
        if df_clean['Fecha'].dtype == 'object':
            # Ya es string, no hacer nada
            pass
        else:
            # 1. Aseguramos que Pandas entienda que son fechas con formato de nuestro país
            df_clean['Fecha'] = pd.to_datetime(df_clean['Fecha'], dayfirst=True, errors='coerce')

            # 2. Ahora sí aplicamos el formato visual de manera segura (esta es tu línea original)
            df_clean['Fecha'] = df_clean['Fecha'].dt.strftime('%d/%m/%Y')
        
        # Calcular totales para las columnas numéricas
        totals_row = {}
        for col in df_clean.columns:
            if col in numeric_cols:
                totals_row[col] = df_clean[col].sum()
            else:
                totals_row[col] = ''
        
        # Agregar fila de totales
        df_clean = pd.concat([df_clean, pd.DataFrame([totals_row])], ignore_index=True)
        
        return df_clean
    
    def save_to_excel(self, output_path: str, df: pd.DataFrame = None):
        """Guarda el DataFrame procesado en un nuevo Excel"""
        if df is None:
            df = self.clean_data()
        
        df.to_excel(output_path, index=False, engine='openpyxl')
        return output_path
    
    def add_sheet_to_workbook(self, workbook_path: str, sheet_name: str, df: pd.DataFrame = None) -> str:
        """
        Agrega una nueva pestaña a un Excel existente
        Returns: "success" o "exists" si la pestaña ya existe
        """
        if df is None:
            df = self.clean_data()
        
        # Abrir el workbook existente
        try:
            wb = openpyxl.load_workbook(workbook_path)
        except FileNotFoundError:
            # Si el archivo no existe, crear uno nuevo
            wb = openpyxl.Workbook()
            # Eliminar hojas por defecto
            for sheet in wb.worksheets:
                wb.remove(sheet)
        
        # Verificar si la pestaña ya existe
        if sheet_name in wb.sheetnames:
            return "exists"
        
        # Eliminar la hoja temporal si existe (creada al crear archivo nuevo)
        if "_temp" in wb.sheetnames:
            temp_sheet = wb["_temp"]
            wb.remove(temp_sheet)
        
        # Crear nueva pestaña
        ws = wb.create_sheet(title=sheet_name)
        
        # Escribir los datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Formato para la primera fila (encabezados)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Formato para la última fila (totales)
                if r_idx == len(df) + 1:
                    cell.font = Font(bold=True)
        
        # Guardar el workbook
        wb.save(workbook_path)
        return "success"


def process_excel_file(file_path: str, output_path: str = None) -> Dict:
    """
    Función helper para procesar un Excel completo
    Returns: dict con info del procesamiento
    """
    processor = ExcelProcessor(file_path)
    processor.read_excel()
    
    month, year = processor.detect_month()
    month_name = processor.get_month_name(month)
    
    df_clean = processor.clean_data()
    
    result = {
        "month": month,
        "year": year,
        "month_name": month_name,
        "rows_processed": len(df_clean) - 1,  # -1 por la fila de totales
        "columns_kept": len(df_clean.columns)
    }
    
    if output_path:
        processor.save_to_excel(output_path, df_clean)
        result["output_path"] = output_path
    
    return result
